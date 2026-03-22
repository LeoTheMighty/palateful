"""Spreadsheet parser — extracts rows from CSV/XLSX as text for LLM parsing."""

from __future__ import annotations

import csv
import io
import logging

logger = logging.getLogger(__name__)

MAX_ROWS = 200


def parse_spreadsheet(file_bytes: bytes, filename: str) -> list[str]:
    """Parse a CSV or XLSX file and return rows as serialized text.

    Each row is serialized as "Header1: Value1\\nHeader2: Value2\\n..."
    which the TextExtractor LLM can parse semantically.

    Returns:
        List of text strings, one per non-empty row (capped at MAX_ROWS).
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext == "csv":
        return _parse_csv(file_bytes)
    elif ext in ("xlsx", "xls"):
        return _parse_xlsx(file_bytes)
    else:
        raise ValueError(f"Unsupported file type: .{ext}")


def _parse_csv(file_bytes: bytes) -> list[str]:
    text = file_bytes.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows_iter = iter(reader)

    try:
        headers = next(rows_iter)
    except StopIteration:
        return []

    results: list[str] = []
    for row in rows_iter:
        if len(results) >= MAX_ROWS:
            break
        if not any(cell.strip() for cell in row):
            continue  # skip empty rows
        serialized = _serialize_row(headers, row)
        if serialized:
            results.append(serialized)

    return results


def _parse_xlsx(file_bytes: bytes) -> list[str]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError("openpyxl is required for XLSX parsing") from e

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []

    headers = [str(h) if h is not None else f"Column{i+1}" for i, h in enumerate(header_row)]

    results: list[str] = []
    for row in rows_iter:
        if len(results) >= MAX_ROWS:
            break
        str_row = [str(cell) if cell is not None else "" for cell in row]
        if not any(cell.strip() for cell in str_row):
            continue
        serialized = _serialize_row(headers, str_row)
        if serialized:
            results.append(serialized)

    wb.close()
    return results


def _serialize_row(headers: list[str], row: list[str]) -> str | None:
    """Serialize a row as 'Header: Value' pairs, skipping empty values."""
    parts = []
    for i, header in enumerate(headers):
        value = row[i].strip() if i < len(row) else ""
        if value:
            parts.append(f"{header}: {value}")
    return "\n".join(parts) if parts else None
